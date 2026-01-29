"""
Alphabot wins exporter (used by Telegram /wins and /winsexcel commands).

Exports an XLSX with columns:
- Project
- Chain
- Mint date (UTC)   (blank if unknown)
- Supply            (blank if unknown)
- Mint price        (blank if unknown)
- Twitter           (blank if unknown)

Notes:
- "picked" in Alphabot API is the **win selection timestamp**, NOT mint date.
- Some projects simply don't publish mint/supply/price; we keep those blank.
- When regenerating, projects whose mint date is older than N days are excluded.
"""

from __future__ import annotations

import asyncio
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_URL = "https://www.alphabot.app"
API_PROJECTS = f"{BASE_URL}/api/projects"
PROJECT_URL = f"{BASE_URL}/_"
COOKIE_NAME = "__Secure-next-auth.session-token"
NEXT_DATA_RE = re.compile(r'<script[^>]+id="__NEXT_DATA__"[^>]*>(.*?)</script>', re.DOTALL)


def _first_non_empty(*vals: Any) -> str:
    for v in vals:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return str(v)
    return ""


def _extract_slug(item: Dict[str, Any]) -> str:
    for k in ("slug", "projectSlug", "project", "handle"):
        v = item.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    pd = item.get("projectData") or {}
    for k in ("slug", "projectSlug", "project", "handle"):
        v = pd.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _ms_to_dt_utc(ts_ms: Any) -> Optional[datetime]:
    if not ts_ms:
        return None
    try:
        return datetime.fromtimestamp(int(ts_ms) / 1000, tz=timezone.utc)
    except Exception:
        return None


def _iso_to_dt_utc(iso: str) -> Optional[datetime]:
    if not iso:
        return None
    try:
        s = iso.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def _dt_to_str_utc(dt: Optional[datetime]) -> str:
    if not dt:
        return ""
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _find_first_key(obj: Any, keys: Tuple[str, ...]) -> Optional[Any]:
    stack = [obj]
    seen = 0
    while stack:
        cur = stack.pop()
        seen += 1
        if seen > 200_000:
            return None
        if isinstance(cur, dict):
            for k, v in cur.items():
                if k in keys:
                    return v
                stack.append(v)
        elif isinstance(cur, list):
            stack.extend(cur)
    return None


async def _with_retries(coro_fn, *, tries: int = 4, base_delay: float = 0.7):
    last = None
    for i in range(tries):
        try:
            return await coro_fn()
        except Exception as e:
            last = e
            await asyncio.sleep(base_delay * (2**i))
    raise last  # type: ignore[misc]


async def _fetch_wins_page(client: httpx.AsyncClient, session_token: str, page_num: int) -> List[Dict[str, Any]]:
    resp = await client.get(
        API_PROJECTS,
        params={
            "sort": "newest",
            "scope": "all",
            "showHidden": "false",
            "pageSize": "100",
            "pageNum": str(page_num),
            "search": "",
            "project": "",
            "includeProject": "true",
            "filter": "winners",
        },
        cookies={COOKIE_NAME: session_token},
        headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
    )
    if resp.status_code in (401, 403):
        raise RuntimeError("Auth failed (401/403). ALPHABOT_SESSION_TOKEN is missing/expired.")
    resp.raise_for_status()
    data = resp.json()
    return data if isinstance(data, list) else []


async def _fetch_next_data(client: httpx.AsyncClient, slug: str) -> Optional[Dict[str, Any]]:
    url = f"{PROJECT_URL}/{slug}"
    resp = await client.get(url, headers={"User-Agent": "Mozilla/5.0", "Accept": "text/html"})
    if resp.status_code != 200:
        return None
    m = NEXT_DATA_RE.search(resp.text)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None


def _parse_project_page(next_data: Dict[str, Any]) -> Tuple[Optional[datetime], str, str, str]:
    """
    Return (mint_dt_utc, supply, price, twitterUrl).
    """
    mint_raw = _find_first_key(next_data, ("mintDate", "mint_date", "mintAt", "mintTime", "mint"))
    supply_raw = _find_first_key(next_data, ("supply", "maxSupply", "totalSupply"))
    price_raw = _find_first_key(next_data, ("wlPrice", "price", "mintPrice"))
    twitter_raw = _find_first_key(next_data, ("twitterUrl", "twitter", "twitter_link"))

    mint_dt = None
    if isinstance(mint_raw, (int, float)) or (isinstance(mint_raw, str) and mint_raw.isdigit()):
        mint_dt = _ms_to_dt_utc(mint_raw)
    elif isinstance(mint_raw, str):
        mint_dt = _iso_to_dt_utc(mint_raw)

    supply = ""
    if isinstance(supply_raw, (int, float)):
        supply = str(int(supply_raw))
    elif isinstance(supply_raw, str):
        supply = supply_raw.strip()

    price = ""
    if isinstance(price_raw, (int, float)):
        price = str(price_raw)
    elif isinstance(price_raw, str):
        price = price_raw.strip()

    twitter = twitter_raw.strip() if isinstance(twitter_raw, str) else ""
    return mint_dt, supply, price, twitter


@dataclass
class WinRow:
    project: str
    chain: str
    mint_dt: Optional[datetime]
    supply: str
    price: str
    twitter: str
    # When the win was picked/selected (from Alphabot "picked" field)
    picked_dt: Optional[datetime] = None


def _write_xlsx(rows: List["WinRow"], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "wins"

    headers = ["Project", "Chain", "Mint date (UTC)", "Supply", "Mint price", "Twitter"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([r.project, r.chain, _dt_to_str_utc(r.mint_dt), r.supply, r.price, r.twitter])

    # Make Twitter clickable
    for row_idx in range(2, len(rows) + 2):
        tw = ws.cell(row=row_idx, column=6)
        if tw.value and str(tw.value).startswith("http"):
            tw.hyperlink = str(tw.value)
            tw.style = "Hyperlink"

    for col_idx in range(1, len(headers) + 1):
        max_len = 10
        for row in range(1, len(rows) + 2):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, min(60, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    ws.freeze_panes = "A2"
    wb.save(out_path)


async def fetch_wins_rows(
    *,
    session_token: str,
    cutoff_date_utc: Optional[datetime] = None,
    enrich_concurrency: int = 6,
    drop_mints_older_than_days: int = 7,
) -> List["WinRow"]:
    """
    Fetch and enrich Alphabot wins and return them as WinRow objects.
    """
    if cutoff_date_utc is None:
        cutoff_date_utc = datetime(2026, 1, 26, 0, 0, 0, tzinfo=timezone.utc)

    cutoff_ms = int(cutoff_date_utc.timestamp() * 1000)
    now = datetime.now(timezone.utc)
    drop_before = now - timedelta(days=drop_mints_older_than_days)
    sem = asyncio.Semaphore(max(1, min(int(enrich_concurrency), 12)))

    limits = httpx.Limits(max_connections=20, max_keepalive_connections=10)
    timeout = httpx.Timeout(30.0)

    async with httpx.AsyncClient(timeout=timeout, limits=limits, follow_redirects=True) as client:
        items: List[Dict[str, Any]] = []
        page = 0
        while True:
            page_items = await _with_retries(lambda: _fetch_wins_page(client, session_token, page))
            if not page_items:
                break

            items.extend([x for x in page_items if int(x.get("picked") or 0) >= cutoff_ms])

            min_picked = min((int(x.get("picked") or 0) for x in page_items), default=0)
            if min_picked and min_picked < cutoff_ms:
                break
            if len(page_items) < 100:
                break
            page += 1

        # Base rows from API
        rows: List[WinRow] = []
        slugs: List[str] = []
        for it in items:
            pd = it.get("projectData") or {}
            project = _first_non_empty(pd.get("name"), it.get("name"), "Unknown")
            chain = _first_non_empty(it.get("blockchain"), pd.get("blockchain"), "")
            slug = _extract_slug(it)

            mint_dt = _ms_to_dt_utc(pd.get("mintDate")) if pd.get("mintDate") else None
            picked_dt = _ms_to_dt_utc(it.get("picked"))
            supply = ""
            price = _first_non_empty(pd.get("wlPrice"), "")
            twitter = _first_non_empty(pd.get("twitterUrl"), it.get("twitterUrl"), "")

            rows.append(
                WinRow(
                    project=project,
                    chain=chain,
                    mint_dt=mint_dt,
                    supply=supply,
                    price=price,
                    twitter=twitter,
                    picked_dt=picked_dt,
                )
            )
            slugs.append(slug)

        async def enrich_one(slug: str) -> Tuple[Optional[datetime], str, str, str]:
            if not slug:
                return None, "", "", ""
            async with sem:
                nd = await _with_retries(lambda: _fetch_next_data(client, slug), tries=3, base_delay=0.6)
                if not nd:
                    return None, "", "", ""
                return _parse_project_page(nd)

        enriched = await asyncio.gather(*(enrich_one(s) for s in slugs), return_exceptions=True)
        for i, extra in enumerate(enriched):
            if isinstance(extra, Exception):
                continue
            mint_dt, supply, price, twitter = extra
            if mint_dt and not rows[i].mint_dt:
                rows[i].mint_dt = mint_dt
            if supply and not rows[i].supply:
                rows[i].supply = supply
            if price and not rows[i].price:
                rows[i].price = price
            if twitter and not rows[i].twitter:
                rows[i].twitter = twitter

        # Drop projects whose mint date is older than N days
        filtered_rows: List[WinRow] = []
        for r in rows:
            if r.mint_dt and r.mint_dt < drop_before:
                continue
            filtered_rows.append(r)

        return filtered_rows


async def export_wins_xlsx(
    *,
    session_token: str,
    out_path: str,
    cutoff_date_utc: Optional[datetime] = None,
    enrich_concurrency: int = 6,
    drop_mints_older_than_days: int = 7,
) -> Tuple[str, int]:
    """
    Returns (out_path, rows_written)
    """
    rows = await fetch_wins_rows(
        session_token=session_token,
        cutoff_date_utc=cutoff_date_utc,
        enrich_concurrency=enrich_concurrency,
        drop_mints_older_than_days=drop_mints_older_than_days,
    )
    _write_xlsx(rows, out_path)
    return out_path, len(rows)

